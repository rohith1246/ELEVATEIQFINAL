"""
Daily Tasks Blueprint Routes.

Handles employee daily task logging, updates, history lookups, admin task roster queries,
day-by-day summary reports, and multi-sheet formatted Excel (.xlsx) export report generation.
"""

import io
import logging
from datetime import datetime, timezone, timedelta
from flask import Blueprint, request, jsonify, send_file
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_connection
from ..auth import get_current_user, require_role
from ..config import safe_error

logger = logging.getLogger(__name__)
IST = timezone(timedelta(hours=5, minutes=30))

daily_tasks_bp = Blueprint("daily_tasks", __name__)


@daily_tasks_bp.route("/api/daily-tasks", methods=["POST"])
def create_daily_task():
    """
    Creates a new daily task entry for the logged-in employee.
    """
    user = get_current_user()
    if not user:
        return jsonify({"error": "Unauthorized: Please log in first"}), 401

    data = request.json or {}
    task_date_str = data.get("task_date")
    project_name = data.get("project_name", "").strip()
    task_description = data.get("task_description", "").strip()
    hours_spent = data.get("hours_spent", 0.0)
    status = data.get("status", "Completed").strip()
    remarks = data.get("remarks", "").strip()

    if not task_description:
        return jsonify({"error": "Task description is required"}), 400

    try:
        hours_val = float(hours_spent) if hours_spent is not None else 0.0
    except (ValueError, TypeError):
        hours_val = 0.0

    if not task_date_str:
        task_date = datetime.now(IST).date()
    else:
        try:
            task_date = datetime.strptime(task_date_str, "%Y-%m-%d").date()
        except ValueError:
            task_date = datetime.now(IST).date()

    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        user_db_id = user.get("user_id") or user.get("id")
        cursor.execute("SELECT id FROM employees WHERE user_id = %s", (user_db_id,))
        emp_row = cursor.fetchone()
        
        if emp_row:
            emp_id = emp_row["id"]
        else:
            emp_id = user_db_id

        cursor.execute(
            """
            INSERT INTO daily_tasks (employee_id, task_date, project_name, task_description, hours_spent, status, remarks)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id, created_at
            """,
            (emp_id, task_date, project_name, task_description, hours_val, status, remarks)
        )
        new_row = cursor.fetchone()
        conn.commit()

        return jsonify({
            "message": "Daily task logged successfully!",
            "id": new_row["id"],
            "task_date": task_date.isoformat()
        }), 201
    except Exception as e:
        conn.rollback()
        logger.error(f"Error creating daily task: {e}")
        return jsonify(safe_error()), 500
    finally:
        cursor.close()
        conn.close()


@daily_tasks_bp.route("/api/daily-tasks/my-tasks", methods=["GET"])
def get_my_daily_tasks():
    """
    Retrieves personal daily tasks submitted by the logged-in employee.
    """
    user = get_current_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    user_db_id = user.get("user_id") or user.get("id")
    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("SELECT id FROM employees WHERE user_id = %s", (user_db_id,))
        emp_row = cursor.fetchone()
        emp_id = emp_row["id"] if emp_row else user_db_id

        cursor.execute(
            """
            SELECT t.*, e.employee_id as employee_code, u.name as employee_name, COALESCE(e.shift, 'Day Shift') as shift
            FROM daily_tasks t
            JOIN employees e ON t.employee_id = e.id
            JOIN users u ON e.user_id = u.id
            WHERE t.employee_id = %s
            ORDER BY t.task_date DESC, t.created_at DESC
            """,
            (emp_id,)
        )
        tasks = cursor.fetchall()
        for t in tasks:
            if t.get("task_date"):
                t["task_date"] = t["task_date"].isoformat()
            if t.get("created_at"):
                t["created_at"] = t["created_at"].isoformat()
            if t.get("hours_spent") is not None:
                t["hours_spent"] = float(t["hours_spent"])

        return jsonify(tasks), 200
    except Exception as e:
        logger.error(f"Error getting my daily tasks: {e}")
        return jsonify(safe_error()), 500
    finally:
        cursor.close()
        conn.close()


@daily_tasks_bp.route("/api/daily-tasks/admin/tasks", methods=["GET"])
def get_admin_daily_tasks():
    """
    Retrieves daily tasks across all employees for Admins and HR.
    Supports start_date, end_date, employee_id, shift, and status filters.
    """
    user = get_current_user()
    if not user or user.get("role", "").lower() not in ("admin", "hr"):
        return jsonify({"error": "Forbidden: Admin or HR access required"}), 403

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    emp_filter = request.args.get("employee_id")
    shift_filter = request.args.get("shift")
    status_filter = request.args.get("status")

    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        query = """
            SELECT t.*, e.employee_id as employee_code, u.name as employee_name, e.department, COALESCE(e.shift, 'Day Shift') as shift
            FROM daily_tasks t
            JOIN employees e ON t.employee_id = e.id
            JOIN users u ON e.user_id = u.id
            WHERE (u.portal = 'elevateiq' OR u.portal = 'both' OR u.portal IS NULL)
        """
        params = []

        if start_date:
            query += " AND t.task_date >= %s"
            params.append(start_date)
        if end_date:
            query += " AND t.task_date <= %s"
            params.append(end_date)
        if emp_filter and emp_filter != "All":
            query += " AND e.id = %s"
            params.append(emp_filter)
        if shift_filter and shift_filter != "All":
            query += " AND LOWER(COALESCE(e.shift, 'Day Shift')) = %s"
            params.append(shift_filter.lower())
        if status_filter and status_filter != "All":
            query += " AND LOWER(t.status) = %s"
            params.append(status_filter.lower())

        query += " ORDER BY t.task_date DESC, t.created_at DESC"

        cursor.execute(query, params)
        tasks = cursor.fetchall()
        for t in tasks:
            if t.get("task_date"):
                t["task_date"] = t["task_date"].isoformat()
            if t.get("created_at"):
                t["created_at"] = t["created_at"].isoformat()
            if t.get("hours_spent") is not None:
                t["hours_spent"] = float(t["hours_spent"])

        return jsonify(tasks), 200
    except Exception as e:
        logger.error(f"Error getting admin daily tasks: {e}")
        return jsonify(safe_error()), 500
    finally:
        cursor.close()
        conn.close()


@daily_tasks_bp.route("/api/daily-tasks/admin/daily-summary", methods=["GET"])
def get_admin_daily_summary():
    """
    Retrieves day-by-day aggregated daily task metrics for Admin reports.
    Groups tasks by task_date and computes total tasks, employees, hours, and shift breakdowns.
    """
    user = get_current_user()
    if not user or user.get("role", "").lower() not in ("admin", "hr"):
        return jsonify({"error": "Forbidden: Admin or HR access required"}), 403

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        query = """
            SELECT 
                t.task_date,
                COUNT(t.id) as total_tasks,
                COUNT(DISTINCT t.employee_id) as total_employees,
                COALESCE(SUM(t.hours_spent), 0.0) as total_hours,
                COUNT(CASE WHEN LOWER(t.status) = 'completed' THEN 1 END) as completed_tasks,
                COUNT(CASE WHEN LOWER(t.status) != 'completed' THEN 1 END) as pending_tasks,
                COALESCE(SUM(CASE WHEN LOWER(COALESCE(e.shift, 'day shift')) = 'day shift' THEN t.hours_spent ELSE 0 END), 0.0) as day_shift_hours,
                COALESCE(SUM(CASE WHEN LOWER(COALESCE(e.shift, 'day shift')) = 'night shift' THEN t.hours_spent ELSE 0 END), 0.0) as night_shift_hours
            FROM daily_tasks t
            JOIN employees e ON t.employee_id = e.id
            JOIN users u ON e.user_id = u.id
            WHERE (u.portal = 'elevateiq' OR u.portal = 'both' OR u.portal IS NULL)
        """
        params = []

        if start_date:
            query += " AND t.task_date >= %s"
            params.append(start_date)
        if end_date:
            query += " AND t.task_date <= %s"
            params.append(end_date)

        query += " GROUP BY t.task_date ORDER BY t.task_date DESC"

        cursor.execute(query, params)
        summaries = cursor.fetchall()
        for s in summaries:
            if s.get("task_date"):
                s["task_date"] = s["task_date"].isoformat()
            s["total_hours"] = float(s["total_hours"])
            s["day_shift_hours"] = float(s["day_shift_hours"])
            s["night_shift_hours"] = float(s["night_shift_hours"])

        return jsonify(summaries), 200
    except Exception as e:
        logger.error(f"Error getting admin daily summary report: {e}")
        return jsonify(safe_error()), 500
    finally:
        cursor.close()
        conn.close()


@daily_tasks_bp.route("/api/daily-tasks/<int:task_id>", methods=["PUT"])
def update_daily_task(task_id):
    """
    Updates an existing daily task entry.
    """
    user = get_current_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json or {}
    project_name = data.get("project_name", "").strip()
    task_description = data.get("task_description", "").strip()
    hours_spent = data.get("hours_spent")
    status = data.get("status", "").strip()
    remarks = data.get("remarks", "").strip()

    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("SELECT * FROM daily_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        if not task:
            return jsonify({"error": "Task log not found"}), 404

        user_role = user.get("role", "").lower()
        user_db_id = user.get("user_id") or user.get("id")

        if user_role not in ("admin", "hr"):
            cursor.execute("SELECT id FROM employees WHERE user_id = %s", (user_db_id,))
            emp_row = cursor.fetchone()
            emp_id = emp_row["id"] if emp_row else user_db_id
            if task["employee_id"] != emp_id:
                return jsonify({"error": "Forbidden: You can only edit your own tasks"}), 403

        query = "UPDATE daily_tasks SET updated_at = CURRENT_TIMESTAMP"
        params = []

        if project_name:
            query += ", project_name = %s"
            params.append(project_name)
        if task_description:
            query += ", task_description = %s"
            params.append(task_description)
        if hours_spent is not None:
            query += ", hours_spent = %s"
            params.append(float(hours_spent))
        if status:
            query += ", status = %s"
            params.append(status)
        if remarks is not None:
            query += ", remarks = %s"
            params.append(remarks)

        query += " WHERE id = %s"
        params.append(task_id)

        cursor.execute(query, params)
        conn.commit()
        return jsonify({"message": "Task log updated successfully"}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"Error updating task: {e}")
        return jsonify(safe_error()), 500
    finally:
        cursor.close()
        conn.close()


@daily_tasks_bp.route("/api/daily-tasks/<int:task_id>", methods=["DELETE"])
def delete_daily_task(task_id):
    """
    Deletes a daily task entry.
    """
    user = get_current_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cursor.execute("SELECT * FROM daily_tasks WHERE id = %s", (task_id,))
        task = cursor.fetchone()
        if not task:
            return jsonify({"error": "Task log not found"}), 404

        user_role = user.get("role", "").lower()
        user_db_id = user.get("user_id") or user.get("id")

        if user_role not in ("admin", "hr"):
            cursor.execute("SELECT id FROM employees WHERE user_id = %s", (user_db_id,))
            emp_row = cursor.fetchone()
            emp_id = emp_row["id"] if emp_row else user_db_id
            if task["employee_id"] != emp_id:
                return jsonify({"error": "Forbidden: You can only delete your own tasks"}), 403

        cursor.execute("DELETE FROM daily_tasks WHERE id = %s", (task_id,))
        conn.commit()
        return jsonify({"message": "Task log deleted successfully"}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"Error deleting task: {e}")
        return jsonify(safe_error()), 500
    finally:
        cursor.close()
        conn.close()


@daily_tasks_bp.route("/api/daily-tasks/export-excel", methods=["GET"])
def export_daily_tasks_excel():
    """
    Generates and downloads a multi-sheet formatted Excel spreadsheet (.xlsx) of daily employee tasks.
    Sheet 1: Detailed Tasks Roster
    Sheet 2: Day-by-Day Aggregated Summary Report
    """
    user = get_current_user()
    if not user or user.get("role", "").lower() not in ("admin", "hr"):
        return jsonify({"error": "Forbidden: Admin or HR access required"}), 403

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    emp_filter = request.args.get("employee_id")
    shift_filter = request.args.get("shift")
    status_filter = request.args.get("status")

    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        # Fetch Detailed Tasks
        query = """
            SELECT t.*, e.employee_id as employee_code, u.name as employee_name, e.department, COALESCE(e.shift, 'Day Shift') as shift
            FROM daily_tasks t
            JOIN employees e ON t.employee_id = e.id
            JOIN users u ON e.user_id = u.id
            WHERE (u.portal = 'elevateiq' OR u.portal = 'both' OR u.portal IS NULL)
        """
        params = []

        if start_date:
            query += " AND t.task_date >= %s"
            params.append(start_date)
        if end_date:
            query += " AND t.task_date <= %s"
            params.append(end_date)
        if emp_filter and emp_filter != "All":
            query += " AND e.id = %s"
            params.append(emp_filter)
        if shift_filter and shift_filter != "All":
            query += " AND LOWER(COALESCE(e.shift, 'Day Shift')) = %s"
            params.append(shift_filter.lower())
        if status_filter and status_filter != "All":
            query += " AND LOWER(t.status) = %s"
            params.append(status_filter.lower())

        query += " ORDER BY t.task_date DESC, u.name ASC"
        cursor.execute(query, params)
        tasks = cursor.fetchall()

        # Fetch Day-Wise Aggregated Summary
        sum_query = """
            SELECT 
                t.task_date,
                COUNT(t.id) as total_tasks,
                COUNT(DISTINCT t.employee_id) as total_employees,
                COALESCE(SUM(t.hours_spent), 0.0) as total_hours,
                COUNT(CASE WHEN LOWER(t.status) = 'completed' THEN 1 END) as completed_tasks,
                COUNT(CASE WHEN LOWER(t.status) != 'completed' THEN 1 END) as pending_tasks,
                COALESCE(SUM(CASE WHEN LOWER(COALESCE(e.shift, 'day shift')) = 'day shift' THEN t.hours_spent ELSE 0 END), 0.0) as day_shift_hours,
                COALESCE(SUM(CASE WHEN LOWER(COALESCE(e.shift, 'day shift')) = 'night shift' THEN t.hours_spent ELSE 0 END), 0.0) as night_shift_hours
            FROM daily_tasks t
            JOIN employees e ON t.employee_id = e.id
            JOIN users u ON e.user_id = u.id
            WHERE (u.portal = 'elevateiq' OR u.portal = 'both' OR u.portal IS NULL)
        """
        sum_params = []
        if start_date:
            sum_query += " AND t.task_date >= %s"
            sum_params.append(start_date)
        if end_date:
            sum_query += " AND t.task_date <= %s"
            sum_params.append(end_date)
        sum_query += " GROUP BY t.task_date ORDER BY t.task_date DESC"

        cursor.execute(sum_query, sum_params)
        day_summaries = cursor.fetchall()

        # Create OpenPyXL Workbook
        wb = openpyxl.Workbook()
        
        # Styles definition
        header_title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
        sub_title_font = Font(name="Calibri", size=10, italic=True, color="475569")
        table_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        status_fills = {
            "completed": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
            "in progress": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
            "pending": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "blocked": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        }
        status_fonts = {
            "completed": Font(name="Calibri", size=10, bold=True, color="065F46"),
            "in progress": Font(name="Calibri", size=10, bold=True, color="1E40AF"),
            "pending": Font(name="Calibri", size=10, bold=True, color="92400E"),
            "blocked": Font(name="Calibri", size=10, bold=True, color="991B1B")
        }

        # -------------------------------------------------------------
        # SHEET 1: DETAILED TASKS ROSTER
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Detailed Tasks Log"
        ws1.views.sheetView[0].showGridLines = True

        ws1.merge_cells("A1:K1")
        ws1["A1"] = "ELEVATEIQ SOFT TECH PRIVATE LIMITED - DAILY TASKS ROSTER"
        ws1["A1"].font = header_title_font
        ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws1.merge_cells("A2:K2")
        ws1["A2"] = f"Export Date: {datetime.now(IST).strftime('%Y-%m-%d %H:%M IST')} | Filter Range: {start_date or 'All Time'} to {end_date or 'Present'}"
        ws1["A2"].font = sub_title_font
        ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")

        ws1.row_dimensions[1].height = 28
        ws1.row_dimensions[2].height = 18

        headers1 = [
            "Task ID", "Date", "Employee ID", "Employee Name", "Shift", 
            "Department", "Project", "Task Description", "Hours Spent", "Status", "Remarks"
        ]
        
        ws1.row_dimensions[4].height = 25
        for col_num, h_text in enumerate(headers1, 1):
            cell = ws1.cell(row=4, column=col_num, value=h_text)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        curr_row = 5
        tot_hrs = 0.0

        for t in tasks:
            h_val = float(t.get("hours_spent") or 0.0)
            tot_hrs += h_val
            st_val = t.get("status") or "Completed"

            row_vals = [
                t["id"], str(t.get("task_date") or "-"), t.get("employee_code") or "-", 
                t.get("employee_name") or "-", t.get("shift") or "Day Shift", 
                t.get("department") or "General", t.get("project_name") or "General Task", 
                t.get("task_description") or "-", h_val, st_val, t.get("remarks") or "-"
            ]

            ws1.row_dimensions[curr_row].height = 22
            for col_num, val in enumerate(row_vals, 1):
                cell = ws1.cell(row=curr_row, column=col_num, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border

                if col_num in (1, 2, 3, 5, 9):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                if col_num == 9:
                    cell.number_format = "0.00"

                if col_num == 10:
                    st_key = st_val.lower()
                    if st_key in status_fills:
                        cell.fill = status_fills[st_key]
                        cell.font = status_fonts[st_key]

            curr_row += 1

        # Summary Row
        curr_row += 1
        ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
        ws1.cell(row=curr_row, column=1, value="TOTAL LOGGED HOURS").font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
        ws1.cell(row=curr_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws1.cell(row=curr_row, column=9, value=tot_hrs).font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
        ws1.cell(row=curr_row, column=9).number_format = "0.00"
        ws1.cell(row=curr_row, column=9).alignment = Alignment(horizontal="center", vertical="center")

        for c in range(1, 12):
            ws1.cell(row=curr_row, column=c).border = thin_border
            ws1.cell(row=curr_row, column=c).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        for col in ws1.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 4: continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len: max_len = len(val_str)
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

        ws1.column_dimensions["H"].width = 40

        # -------------------------------------------------------------
        # SHEET 2: DAY-BY-DAY SUMMARY REPORT
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Day Wise Summary")
        ws2.views.sheetView[0].showGridLines = True

        ws2.merge_cells("A1:H1")
        ws2["A1"] = "ELEVATEIQ - DAY-BY-DAY EMPLOYEE TASK SUMMARY REPORT"
        ws2["A1"].font = header_title_font
        ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")

        ws2.merge_cells("A2:H2")
        ws2["A2"] = f"Generated On: {datetime.now(IST).strftime('%Y-%m-%d %H:%M IST')}"
        ws2["A2"].font = sub_title_font
        ws2["A2"].alignment = Alignment(horizontal="left", vertical="center")

        ws2.row_dimensions[1].height = 28
        ws2.row_dimensions[2].height = 18

        headers2 = [
            "Date", "Employees Logged", "Total Tasks Logged", "Total Hours Worked",
            "Completed Tasks", "In Progress / Pending", "Day Shift Hours", "Night Shift Hours"
        ]

        ws2.row_dimensions[4].height = 25
        for col_num, h_text in enumerate(headers2, 1):
            cell = ws2.cell(row=4, column=col_num, value=h_text)
            cell.font = table_header_font
            cell.fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Teal header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        curr_row2 = 5
        tot_sum_hrs = 0.0
        tot_sum_tasks = 0

        for ds in day_summaries:
            d_date = str(ds.get("task_date") or "-")
            emp_cnt = ds.get("total_employees") or 0
            task_cnt = ds.get("total_tasks") or 0
            hrs_cnt = float(ds.get("total_hours") or 0.0)
            comp_cnt = ds.get("completed_tasks") or 0
            pend_cnt = ds.get("pending_tasks") or 0
            day_hrs = float(ds.get("day_shift_hours") or 0.0)
            night_hrs = float(ds.get("night_shift_hours") or 0.0)

            tot_sum_hrs += hrs_cnt
            tot_sum_tasks += task_cnt

            row_vals2 = [d_date, emp_cnt, task_cnt, hrs_cnt, comp_cnt, pend_cnt, day_hrs, night_hrs]

            ws2.row_dimensions[curr_row2].height = 22
            for col_num, val in enumerate(row_vals2, 1):
                cell = ws2.cell(row=curr_row2, column=col_num, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if col_num in (4, 7, 8):
                    cell.number_format = "0.00"

            curr_row2 += 1

        for col in ws2.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 4: continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len: max_len = len(val_str)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ElevateIQ_Daily_Tasks_Report_{datetime.now(IST).strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error exporting daily tasks excel: {e}")
        return jsonify(safe_error()), 500
    finally:
        cursor.close()
        conn.close()
